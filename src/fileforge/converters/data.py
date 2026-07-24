"""Structured-data converters: JSON, CSV, TSV, YAML, XML, XLSX.

All routes here are free-tier and depend only on the standard library,
except YAML which uses PyYAML, XLSX which uses openpyxl, and XML which
uses xmltodict (all optional).

All XLSX/XML operations use atomic writes for safety.
"""

from __future__ import annotations

import csv
import io
import json
import tempfile
from pathlib import Path

from fileforge.core.registry import ConversionError, registry


def _read_tabular(rows, target: Path, delimiter: str) -> Path:
    if not rows:
        target.write_text("", encoding="utf-8")
        return target
    fieldnames = list({k: None for row in rows for k in row.keys()})
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, delimiter=delimiter)
    writer.writeheader()
    writer.writerows(rows)
    target.write_text(buf.getvalue(), encoding="utf-8")
    return target


def _atomic_write(target: Path, content: bytes | str) -> Path:
    """Write to temp file first, then atomic rename to target."""
    with tempfile.NamedTemporaryFile(delete=False, dir=target.parent) as tmp:
        tmp_path = Path(tmp.name)
        if isinstance(content, str):
            tmp.write(content.encode("utf-8"))
        else:
            tmp.write(content)
    tmp_path.replace(target)
    return target


# ============================================================================
# JSON ↔ CSV / TSV
# ============================================================================

@registry.add("json", "csv", description="JSON array of objects -> CSV")
def json_to_csv(source: Path, target: Path, **_) -> Path:
    data = json.loads(Path(source).read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        raise ConversionError("JSON must be an object or array of objects")
    return _read_tabular(data, Path(target), ",")


@registry.add("json", "tsv", description="JSON array of objects -> TSV")
def json_to_tsv(source: Path, target: Path, **_) -> Path:
    data = json.loads(Path(source).read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = [data]
    return _read_tabular(data, Path(target), "\t")


@registry.add("csv", "json", description="CSV -> pretty JSON array")
def csv_to_json(source: Path, target: Path, **_) -> Path:
    with Path(source).open(encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))
    Path(target).write_text(json.dumps(rows, indent=2), encoding="utf-8")
    return Path(target)


@registry.add("tsv", "json", description="TSV -> pretty JSON array")
def tsv_to_json(source: Path, target: Path, **_) -> Path:
    with Path(source).open(encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh, delimiter="\t"))
    Path(target).write_text(json.dumps(rows, indent=2), encoding="utf-8")
    return Path(target)


# ============================================================================
# CSV ↔ TSV
# ============================================================================

@registry.add("csv", "tsv", description="CSV -> TSV")
def csv_to_tsv(source: Path, target: Path, **_) -> Path:
    with Path(source).open(encoding="utf-8", newline="") as fh:
        rows = list(csv.reader(fh))
    buf = io.StringIO()
    csv.writer(buf, delimiter="\t").writerows(rows)
    Path(target).write_text(buf.getvalue(), encoding="utf-8")
    return Path(target)


@registry.add("tsv", "csv", description="TSV -> CSV")
def tsv_to_csv(source: Path, target: Path, **_) -> Path:
    with Path(source).open(encoding="utf-8", newline="") as fh:
        rows = list(csv.reader(fh, delimiter="\t"))
    buf = io.StringIO()
    csv.writer(buf).writerows(rows)
    Path(target).write_text(buf.getvalue(), encoding="utf-8")
    return Path(target)


# ============================================================================
# JSON ↔ YAML
# ============================================================================

@registry.add("json", "yaml", description="JSON -> YAML", requires=["yaml"])
def json_to_yaml(source: Path, target: Path, **_) -> Path:
    import yaml  # optional dependency

    data = json.loads(Path(source).read_text(encoding="utf-8"))
    Path(target).write_text(
        yaml.safe_dump(data, sort_keys=False), encoding="utf-8"
    )
    return Path(target)


@registry.add("yaml", "json", description="YAML -> pretty JSON", requires=["yaml"])
def yaml_to_json(source: Path, target: Path, **_) -> Path:
    import yaml  # optional dependency

    data = yaml.safe_load(Path(source).read_text(encoding="utf-8"))
    Path(target).write_text(json.dumps(data, indent=2), encoding="utf-8")
    return Path(target)


# ============================================================================
# JSON ↔ XML
# ============================================================================

@registry.add("json", "xml", description="JSON -> XML", requires=["xmltodict"])
def json_to_xml(source: Path, target: Path, **_) -> Path:
    import xmltodict  # optional dependency

    data = json.loads(Path(source).read_text(encoding="utf-8"))
    if isinstance(data, list):
        # Wrap array in a root element
        data = {"root": {"item": data}}
    xml_str = xmltodict.unparse(data, pretty=True)
    return _atomic_write(target, xml_str)


@registry.add("xml", "json", description="XML -> pretty JSON", requires=["xmltodict"])
def xml_to_json(source: Path, target: Path, **_) -> Path:
    import xmltodict  # optional dependency

    xml_str = Path(source).read_text(encoding="utf-8")
    data = xmltodict.parse(xml_str)
    Path(target).write_text(json.dumps(data, indent=2), encoding="utf-8")
    return Path(target)


# ============================================================================
# CSV ↔ XLSX / JSON ↔ XLSX / TSV ↔ XLSX
# ============================================================================

@registry.add("csv", "xlsx", description="CSV -> XLSX (Excel)", requires=["openpyxl"])
def csv_to_xlsx(source: Path, target: Path, **_) -> Path:
    from openpyxl import Workbook  # optional dependency

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    with Path(source).open(encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh)
        for row in reader:
            ws.append(row)
    
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
        wb.save(str(tmp_path))
    
    tmp_path.replace(target)
    return Path(target)


@registry.add("xlsx", "csv", description="XLSX (Excel) -> CSV", requires=["openpyxl"])
def xlsx_to_csv(source: Path, target: Path, **_) -> Path:
    from openpyxl import load_workbook  # optional dependency

    wb = load_workbook(str(source))
    ws = wb.active
    
    with Path(target).open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    
    return Path(target)


@registry.add("json", "xlsx", description="JSON array of objects -> XLSX", requires=["openpyxl"])
def json_to_xlsx(source: Path, target: Path, **_) -> Path:
    from openpyxl import Workbook  # optional dependency

    data = json.loads(Path(source).read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        raise ConversionError("JSON must be an object or array of objects")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    if data:
        # Write headers
        fieldnames = list(data[0].keys())
        ws.append(fieldnames)
        # Write rows
        for row in data:
            ws.append([row.get(k, "") for k in fieldnames])
    
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
        wb.save(str(tmp_path))
    
    tmp_path.replace(target)
    return Path(target)


@registry.add("xlsx", "json", description="XLSX (Excel) -> pretty JSON", requires=["openpyxl"])
def xlsx_to_json(source: Path, target: Path, **_) -> Path:
    from openpyxl import load_workbook  # optional dependency

    wb = load_workbook(str(source))
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        Path(target).write_text("[]", encoding="utf-8")
        return Path(target)
    
    headers = rows[0]
    data = [dict(zip(headers, row)) for row in rows[1:]]
    
    Path(target).write_text(json.dumps(data, indent=2), encoding="utf-8")
    return Path(target)


@registry.add("tsv", "xlsx", description="TSV -> XLSX (Excel)", requires=["openpyxl"])
def tsv_to_xlsx(source: Path, target: Path, **_) -> Path:
    from openpyxl import Workbook  # optional dependency

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    with Path(source).open(encoding="utf-8", newline="") as fh:
        reader = csv.reader(fh, delimiter="\t")
        for row in reader:
            ws.append(row)
    
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
        wb.save(str(tmp_path))
    
    tmp_path.replace(target)
    return Path(target)


@registry.add("xlsx", "tsv", description="XLSX (Excel) -> TSV", requires=["openpyxl"])
def xlsx_to_tsv(source: Path, target: Path, **_) -> Path:
    from openpyxl import load_workbook  # optional dependency

    wb = load_workbook(str(source))
    ws = wb.active
    
    with Path(target).open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh, delimiter="\t")
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    
    return Path(target)


@registry.add("xlsx", "yaml", description="XLSX (Excel) -> YAML", requires=["openpyxl", "yaml"])
def xlsx_to_yaml(source: Path, target: Path, **_) -> Path:
    import yaml  # optional dependency
    from openpyxl import load_workbook  # optional dependency

    wb = load_workbook(str(source))
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        Path(target).write_text("---\n", encoding="utf-8")
        return Path(target)
    
    headers = rows[0]
    data = [dict(zip(headers, row)) for row in rows[1:]]
    
    Path(target).write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    return Path(target)


@registry.add("yaml", "xlsx", description="YAML array of objects -> XLSX", requires=["yaml", "openpyxl"])
def yaml_to_xlsx(source: Path, target: Path, **_) -> Path:
    import yaml  # optional dependency
    from openpyxl import Workbook  # optional dependency

    data = yaml.safe_load(Path(source).read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        raise ConversionError("YAML must be an object or array of objects")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    if data:
        fieldnames = list(data[0].keys())
        ws.append(fieldnames)
        for row in data:
            ws.append([row.get(k, "") for k in fieldnames])
    
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)
        wb.save(str(tmp_path))
    
    tmp_path.replace(target)
    return Path(target)

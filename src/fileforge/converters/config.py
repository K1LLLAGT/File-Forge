"""Config format converters: JSON, YAML, TOML, XML.

Handles configuration file conversions with safe parsing and pretty output.
All routes are free-tier except TOML which requires tomli/tomli_write.
"""

from __future__ import annotations

import json
import tempfile
from pathlib import Path

from fileforge.core.registry import ConversionError, registry


# ============================================================================
# JSON ↔ TOML
# ============================================================================

@registry.add("json", "toml", description="JSON -> TOML", requires=["tomli_w"])
def json_to_toml(source: Path, target: Path, **_) -> Path:
    import tomli_w  # optional dependency

    data = json.loads(Path(source).read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ConversionError("JSON must be an object (not array) for TOML")
    
    # Atomic write: write to temp file first
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".toml") as tmp:
        tmp_path = Path(tmp.name)
        tmp.write(tomli_w.dumps(data).encode("utf-8"))
    
    # Atomic rename
    tmp_path.replace(target)
    return Path(target)


@registry.add("toml", "json", description="TOML -> pretty JSON", requires=["tomli"])
def toml_to_json(source: Path, target: Path, **_) -> Path:
    import tomli  # optional dependency

    toml_str = Path(source).read_text(encoding="utf-8")
    data = tomli.loads(toml_str)
    Path(target).write_text(json.dumps(data, indent=2), encoding="utf-8")
    return Path(target)


# ============================================================================
# YAML ↔ TOML
# ============================================================================

@registry.add("yaml", "toml", description="YAML -> TOML", requires=["yaml", "tomli_w"])
def yaml_to_toml(source: Path, target: Path, **_) -> Path:
    import tomli_w  # optional dependency
    import yaml  # optional dependency

    data = yaml.safe_load(Path(source).read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ConversionError("YAML must be an object (not array) for TOML")
    
    with tempfile.NamedTemporaryFile(mode="wb", delete=False, suffix=".toml") as tmp:
        tmp_path = Path(tmp.name)
        tmp.write(tomli_w.dumps(data).encode("utf-8"))
    
    tmp_path.replace(target)
    return Path(target)


@registry.add("toml", "yaml", description="TOML -> YAML", requires=["tomli", "yaml"])
def toml_to_yaml(source: Path, target: Path, **_) -> Path:
    import tomli  # optional dependency
    import yaml  # optional dependency

    toml_str = Path(source).read_text(encoding="utf-8")
    data = tomli.loads(toml_str)
    Path(target).write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    return Path(target)


# ============================================================================
# XLSX ↔ YAML
# ============================================================================

@registry.add("xlsx", "yaml", description="XLSX (Excel) -> YAML", requires=["openpyxl", "yaml"])
def xlsx_to_yaml(source: Path, target: Path, **_) -> Path:
    import yaml  # optional dependency
    from openpyxl import load_workbook  # optional dependency

    wb = load_workbook(str(source))
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        Path(target).write_text("---\n", encoding="utf-8")
        return Path(target)
    
    headers = rows[0]
    data = [dict(zip(headers, row)) for row in rows[1:]]
    
    Path(target).write_text(yaml.safe_dump(data, sort_keys=False), encoding="utf-8")
    return Path(target)


@registry.add("yaml", "xlsx", description="YAML array of objects -> XLSX", requires=["yaml", "openpyxl"])
def yaml_to_xlsx(source: Path, target: Path, **_) -> Path:
    import yaml  # optional dependency
    from openpyxl import Workbook  # optional dependency

    data = yaml.safe_load(Path(source).read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        raise ConversionError("YAML must be an object or array of objects")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    if data:
        fieldnames = list(data[0].keys())
        ws.append(fieldnames)
        for row in data:
            ws.append([row.get(k, "") for k in fieldnames])
    
    wb.save(str(target))
    return Path(target)
